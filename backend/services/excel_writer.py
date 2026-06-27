"""Populate the Solar Excel template with extracted bill data."""

import logging
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from config import EXCEL_TEMPLATE_PATH, OUTPUTS_DIR
from services.bill_validator import parse_numeric
from services.cell_mapping import (
    AUTOMATION_FIELDS,
    CALCULATOR_SHEET_NAME,
    CELL_MAPPING,
    EXTRA_CELL_MAPPING,
    FIELD_BY_NAME,
    FORMULA_CELLS,
)

logger = logging.getLogger(__name__)

# Cells that must never receive written values
_PROTECTED_FORMULA_CELLS = {spec["cell"] for spec in FORMULA_CELLS}


# Fields written to Excel — excludes monthly_history, meter_serial_number, etc.
_EXCEL_FIELD_NAMES = {f["field_name"] for f in AUTOMATION_FIELDS}


def _coerce_cell_value(value: Any, data_type: str) -> str | int | float | None:
    """Convert validated values to types openpyxl writes; numbers stay numeric."""
    if value is None:
        return None

    if data_type in ("Number", "Currency"):
        if isinstance(value, (int, float)):
            return value
        return parse_numeric(value)

    if data_type == "Date":
        return None if value is None else str(value).strip()

    if isinstance(value, (int, float, bool)):
        return value

    text = str(value).strip()
    return text if text else None


def fill_excel_template(data: dict[str, Any]) -> tuple[str, Path]:
    """
    Load the Excel template, write input values to mapped cells, save output.

    Formulas in the template are preserved — only mapped input cells receive values.

    Args:
        data: Extracted (and optionally user-edited) bill field dictionary.

    Returns:
        Tuple of (output_filename, output_path).

    Raises:
        FileNotFoundError: Template file missing.
        ValueError: Workbook cannot be processed.
    """
    if not EXCEL_TEMPLATE_PATH.exists():
        raise FileNotFoundError(
            f"Excel template not found at {EXCEL_TEMPLATE_PATH}. "
            "Run scripts/create_template.py to generate it."
        )

    logger.info("Loading Excel template: %s", EXCEL_TEMPLATE_PATH.name)

    try:
        workbook = load_workbook(EXCEL_TEMPLATE_PATH)
    except Exception as exc:
        raise ValueError(f"Failed to load Excel template: {exc}") from exc

    if CALCULATOR_SHEET_NAME not in workbook.sheetnames:
        raise ValueError(
            f"Sheet '{CALCULATOR_SHEET_NAME}' not found. "
            "Regenerate template with scripts/create_template.py"
        )

    sheet = workbook[CALCULATOR_SHEET_NAME]
    combined_mapping = {**CELL_MAPPING, **EXTRA_CELL_MAPPING}
    filled_count = 0

    for field_name, cell_ref in combined_mapping.items():
        if field_name not in data:
            continue
        if field_name not in _EXCEL_FIELD_NAMES:
            continue

        if cell_ref in _PROTECTED_FORMULA_CELLS:
            logger.warning("Skipping %s: %s is a formula cell", field_name, cell_ref)
            continue

        spec = FIELD_BY_NAME.get(field_name)
        data_type = spec["data_type"] if spec else "Text"
        value = _coerce_cell_value(data[field_name], data_type)
        if value is None:
            continue

        cell = sheet[cell_ref]

        if cell.data_type == "f" and cell.value is not None:
            logger.warning(
                "Skipping %s (%s): cell contains a formula", field_name, cell_ref
            )
            continue

        cell.value = value
        if spec:
            cell.number_format = spec["number_format"]

        filled_count += 1
        logger.debug("Wrote %s -> %s = %s", field_name, cell_ref, value)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_filename = f"Solar_Output_{timestamp}.xlsx"
    output_path = OUTPUTS_DIR / output_filename

    workbook.save(output_path)
    logger.info(
        "Excel saved: %s (%d fields populated)", output_filename, filled_count
    )

    return output_filename, output_path
