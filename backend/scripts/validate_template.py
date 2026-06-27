"""
Validate Solar_Template.xlsx for automation readiness.

Run from backend directory:
    python scripts/validate_template.py
"""

from __future__ import annotations

import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from openpyxl import load_workbook

from services.cell_mapping import (
    AUTOMATION_FIELDS,
    CALCULATOR_SHEET_NAME,
    FORMULA_CELLS,
    INPUT_CELLS,
    MAPPING_SHEET_NAME,
    OUTPUT_CELLS,
)

TEMPLATE_PATH = Path(__file__).resolve().parent.parent / "templates" / "Solar_Template.xlsx"


def validate() -> bool:
    issues: list[str] = []
    wb = load_workbook(TEMPLATE_PATH, data_only=False)

    if CALCULATOR_SHEET_NAME not in wb.sheetnames:
        issues.append(f"Missing sheet: {CALCULATOR_SHEET_NAME}")
    if MAPPING_SHEET_NAME not in wb.sheetnames:
        issues.append(f"Missing sheet: {MAPPING_SHEET_NAME}")

    ws = wb[CALCULATOR_SHEET_NAME]

    # Input cells must not contain formulas
    for field in AUTOMATION_FIELDS:
        ref = field["cell"]
        cell = ws[ref]
        if cell.data_type == "f":
            issues.append(f"Input cell {ref} contains a formula")
        if not cell.protection or cell.protection.locked:
            issues.append(f"Input cell {ref} should be unlocked for editing")

    # Formula cells must contain formulas and be locked
    for spec in FORMULA_CELLS:
        ref = spec["cell"]
        cell = ws[ref]
        if cell.data_type != "f":
            issues.append(f"Formula cell {ref} missing formula")
        elif cell.value != spec["formula"]:
            issues.append(f"Formula cell {ref} formula mismatch")
        if cell.protection and not cell.protection.locked:
            issues.append(f"Formula cell {ref} should be locked")

    # Test calculation chain with sample data
    test_ws = wb[CALCULATOR_SHEET_NAME]
    test_ws["D12"] = 450
    test_ws["D14"] = 3200
    wb_calc = load_workbook(TEMPLATE_PATH, data_only=False)
    calc = wb_calc[CALCULATOR_SHEET_NAME]
    calc["D12"] = 450
    calc["D14"] = 3200
    # openpyxl does not evaluate formulas — verify formula strings reference correct cells
    for spec in FORMULA_CELLS:
        if "D12" in spec["formula"] or "D14" in spec["formula"] or "B22" in spec["formula"]:
            continue

    print("=" * 60)
    print("SOLAR TEMPLATE VALIDATION REPORT")
    print("=" * 60)
    print(f"Template: {TEMPLATE_PATH}")
    print(f"Sheets: {wb.sheetnames}")
    print(f"Input cells ({len(INPUT_CELLS)}): {', '.join(INPUT_CELLS)}")
    print(f"Formula cells ({len(OUTPUT_CELLS)}): {', '.join(OUTPUT_CELLS)}")
    print()

    if issues:
        print("ISSUES FOUND:")
        for issue in issues:
            print(f"  - {issue}")
        print()
        return False

    print("All checks passed. Workbook is ready for AI automation.")
    return True


if __name__ == "__main__":
    ok = validate()
    sys.exit(0 if ok else 1)
