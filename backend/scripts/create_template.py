"""
Create / rebuild the production-ready Solar Load Calculator Excel template.

Features:
  - Automation_Mapping reference sheet
  - Yellow-highlighted AI input cells
  - Green-highlighted formula output cells
  - Sheet protection (formulas locked, inputs editable)
  - Cell comments and number formats
  - openpyxl-compatible structure

Run from backend directory:
    python scripts/create_template.py
"""

from __future__ import annotations

import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side

from services.cell_mapping import (
    AUTOMATION_FIELDS,
    CALCULATOR_SHEET_NAME,
    FORMULA_CELLS,
    MAPPING_SHEET_NAME,
)

TEMPLATE_PATH = Path(__file__).resolve().parent.parent / "templates" / "Solar_Template.xlsx"

FILL_INPUT = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
FILL_FORMULA = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
FILL_HEADER = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
FILL_LABEL = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
FILL_MAPPING_HEADER = PatternFill(start_color="374151", end_color="374151", fill_type="solid")

FONT_HEADER = Font(bold=True, size=12, color="FFFFFF")
FONT_TITLE = Font(bold=True, size=16, color="1E40AF")
FONT_LABEL = Font(bold=True, size=10)
FONT_FORMULA = Font(bold=True, color="166534")
FONT_NORMAL = Font(size=10)

THIN_BORDER = Border(
    left=Side(style="thin", color="E2E8F0"),
    right=Side(style="thin", color="E2E8F0"),
    top=Side(style="thin", color="E2E8F0"),
    bottom=Side(style="thin", color="E2E8F0"),
)

UNLOCKED = Protection(locked=False)
LOCKED = Protection(locked=True)


def _apply_border(ws, cell_ref: str) -> None:
    ws[cell_ref].border = THIN_BORDER


def _build_calculator_sheet(wb: Workbook) -> None:
    ws = wb.active
    ws.title = CALCULATOR_SHEET_NAME

    ws["A1"] = "Energybae AI — Solar Load Calculator"
    ws["A1"].font = FONT_TITLE
    ws.merge_cells("A1:E1")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws["A3"] = "Consumer Details"
    ws["A3"].font = FONT_HEADER
    ws["A3"].fill = FILL_HEADER
    ws["A3"].alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("A3:B3")

    ws["C3"] = "Load & Technical Details"
    ws["C3"].font = FONT_HEADER
    ws["C3"].fill = FILL_HEADER
    ws["C3"].alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("C3:E3")
    ws.row_dimensions[3].height = 22

    label_map = {
        "A4": "Consumer Name:",
        "A5": "Consumer Number:",
        "A6": "Billing Month:",
        "A7": "Billing Period:",
        "A8": "Bill Date:",
        "A9": "Due Date:",
        "A10": "Address:",
        "A11": "Tariff Category:",
        "A12": "Meter Number:",
        "A13": "Division:",
        "A14": "Sub Division:",
        "A15": "Tariff Code:",
        "C4": "Contract Load (kW):",
        "C5": "Connected Load (kW):",
        "C6": "Sanctioned Load (kW):",
        "C7": "Voltage:",
        "C8": "Power Factor:",
        "C9": "Billing Cycle:",
        "C10": "GST Number:",
        "C11": "Previous Reading:",
        "C12": "Units Consumed (kWh):",
        "C13": "Current Reading:",
        "C14": "Bill Amount (₹):",
        "C15": "Fixed Charges:",
        "C16": "Energy Charges:",
        "C17": "Fuel Adjustment:",
        "C18": "Electricity Duty:",
        "C19": "Tax on Sale:",
    }

    for ref, text in label_map.items():
        cell = ws[ref]
        cell.value = text
        cell.font = FONT_LABEL
        cell.fill = FILL_LABEL
        cell.alignment = Alignment(horizontal="right", vertical="center")
        cell.protection = LOCKED
        _apply_border(ws, ref)

    for field in AUTOMATION_FIELDS:
        ref = field["cell"]
        cell = ws[ref]
        cell.value = None
        cell.fill = FILL_INPUT
        cell.number_format = field["number_format"]
        cell.protection = UNLOCKED
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.comment = Comment(field["comment"], "Energybae AI")
        _apply_border(ws, ref)

    ws["A21"] = "Solar Calculation Results"
    ws["A21"].font = FONT_HEADER
    ws["A21"].fill = FILL_HEADER
    ws["A21"].alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("A21:E21")
    ws.row_dimensions[21].height = 22

    output_formats = {
        "B22": '#,##0" kWh"',
        "B23": '0.00" kW"',
        "B24": '₹#,##0.00',
        "B25": '0.0" yrs"',
        "B26": '0.00" t/year"',
    }

    for spec in FORMULA_CELLS:
        label_ref = f"A{spec['cell'][1:]}"
        ws[label_ref] = spec["label"]
        ws[label_ref].font = FONT_LABEL
        ws[label_ref].fill = FILL_LABEL
        ws[label_ref].alignment = Alignment(horizontal="right", vertical="center")
        ws[label_ref].protection = LOCKED
        _apply_border(ws, label_ref)

        out = ws[spec["cell"]]
        out.value = spec["formula"]
        out.fill = FILL_FORMULA
        out.font = FONT_FORMULA
        out.protection = LOCKED
        out.number_format = output_formats.get(spec["cell"], "General")
        out.alignment = Alignment(horizontal="left", vertical="center")
        out.comment = Comment(
            f"{spec['label']}\nFormula cell — do not edit.\n{spec['description']}",
            "Energybae AI",
        )
        _apply_border(ws, spec["cell"])

    for col, width in {"A": 30, "B": 32, "C": 30, "D": 22, "E": 4}.items():
        ws.column_dimensions[col].width = width

    ws.protection.sheet = True
    ws.protection.enable()
    ws.protection.selectLockedCells = True
    ws.protection.selectUnlockedCells = True


def _build_mapping_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet(MAPPING_SHEET_NAME)

    headers = ["Field Name", "Cell Address", "Data Type", "Required", "Description", "Number Format"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = FILL_MAPPING_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx, field in enumerate(AUTOMATION_FIELDS, start=2):
        ws.cell(row=row_idx, column=1, value=field["field_name"])
        ws.cell(row=row_idx, column=2, value=field["cell"])
        ws.cell(row=row_idx, column=3, value=field["data_type"])
        ws.cell(row=row_idx, column=4, value=field["required"])
        ws.cell(row=row_idx, column=5, value=field["description"])
        ws.cell(row=row_idx, column=6, value=field["number_format"])
        for col in range(1, 7):
            ws.cell(row=row_idx, column=col).font = FONT_NORMAL
            ws.cell(row=row_idx, column=col).alignment = Alignment(
                vertical="center", wrap_text=(col == 5)
            )

    start = len(AUTOMATION_FIELDS) + 4
    ws.cell(row=start, column=1, value="FORMULA CELLS (DO NOT WRITE)").font = Font(
        bold=True, size=11, color="166534"
    )
    ws.merge_cells(start_row=start, start_column=1, end_row=start, end_column=6)

    fh_row = start + 1
    for col, h in enumerate(["Cell Address", "Label", "Formula", "Description"], start=1):
        c = ws.cell(row=fh_row, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = FILL_MAPPING_HEADER

    for i, spec in enumerate(FORMULA_CELLS, start=fh_row + 1):
        ws.cell(row=i, column=1, value=spec["cell"])
        ws.cell(row=i, column=2, value=spec["label"])
        ws.cell(row=i, column=3, value=spec["formula"])
        ws.cell(row=i, column=4, value=spec["description"])

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 48
    ws.column_dimensions["F"].width = 18
    ws.freeze_panes = "A2"


def create_template() -> Path:
    wb = Workbook()
    _build_calculator_sheet(wb)
    _build_mapping_sheet(wb)
    TEMPLATE_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(TEMPLATE_PATH)
    return TEMPLATE_PATH


if __name__ == "__main__":
    path = create_template()
    print(f"Template created: {path}")
    print(f"  Input cells:   {len(AUTOMATION_FIELDS)}")
    print(f"  Formula cells: {len(FORMULA_CELLS)}")
    print(f"  Sheets: {CALCULATOR_SHEET_NAME}, {MAPPING_SHEET_NAME}")
